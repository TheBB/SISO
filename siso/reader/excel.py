from pathlib import Path

from typing import Iterable, Tuple
from ..typing import StepData

import numpy as np
import openpyxl

from .. import util as util
from .reader import Reader
from ..geometry import StructuredTopology, Quad, Patch
from ..fields import Field, SimpleField, FieldPatches, Geometry
from ..coords import UTM


def extend(pts: np.ndarray) -> np.ndarray:
    retval = np.zeros((pts.size + 1,), dtype=pts.dtype)
    retval[1:-1] = (pts[:-1] + pts[1:]) / 2
    retval[0] = retval[1] - (pts[1] - pts[0])
    retval[-1] = retval[-2] + (pts[-1] - pts[-2])
    return retval


def safe_float(input):
    try:
        return float(input)
    except ValueError:
        return np.nan


class ExcelGeometryField(SimpleField):

    cells = False
    ncomps = 2
    decompose = False

    reader: 'ExcelReader'

    def __init__(self, reader: 'ExcelReader'):
        self.reader = reader
        self.name = 'geometry'

        coordstring = reader.wb['Maps']['B4'].value
        zone = int(coordstring.split('(')[-1].split(',')[0].split()[1])
        self.fieldtype = Geometry(UTM(f'{zone}N'))

    def patches(self, stepid: int, force: bool = False, **_) -> FieldPatches:
        sheetname = f'V{stepid+1}'
        sheet = self.reader.wb[sheetname]

        xs = np.array([[safe_float(cell.value) for cell in row] for row in sheet.iter_rows(min_col=2, max_row=1)]).reshape(-1)
        ys = np.array([[safe_float(cell.value) for cell in row] for row in sheet.iter_rows(min_row=2, max_col=1)]).reshape(-1)
        xs = extend(xs)
        ys = extend(ys)
        xs, ys = np.meshgrid(xs, ys, indexing='xy')
        nodes = util.flatten_2d(np.array([xs, ys]).T)

        yield self.reader.patch_at(stepid), nodes


class ExcelDataField(SimpleField):

    cells = True
    ncomps = 1
    decompose = False

    reader: 'ExcelReader'

    def __init__(self, reader: 'ExcelReader'):
        self.reader = reader
        self.name = 'V'

    def patches(self, stepid: int, force: bool = False, **_) -> FieldPatches:
        sheetname = f'V{stepid+1}'
        sheet = self.reader.wb[sheetname]

        data = np.array([[safe_float(cell.value) for cell in row] for row in sheet.iter_rows(min_col=2, min_row=2)]).T
        yield self.reader.patch_at(stepid), data.reshape(-1, 1)



class ExcelReader(Reader):

    reader_name = "Excel"

    filename: Path
    wb: openpyxl.Workbook

    @classmethod
    def applicable(cls, filename: Path) -> bool:
        try:
            wb = openpyxl.load_workbook(filename)
            assert 'Info' in wb
        except:
            return False
        return True

    def __init__(self, filename: Path):
        self.filename = filename

    def __enter__(self):
        self.wb = openpyxl.load_workbook(self.filename, 'r')
        return self

    def __exit__(self, *args):
        pass

    def patch_at(self, stepid: int) -> Patch:
        sheetname = f'V{stepid+1}'
        sheet = self.wb[sheetname]
        shape = (sheet.max_column - 1, sheet.max_row - 1)
        return Patch(('geometry',), StructuredTopology(shape, celltype=Quad()))

    def steps(self) -> Iterable[Tuple[int, StepData]]:
        """Iterate over all steps with associated data."""
        nvs = sum(1 for name in self.wb.sheetnames if name.startswith('V') and name[1:].isdigit())
        for stepid in range(nvs):
            yield stepid, {'time': float(stepid)}

    def fields(self) -> Iterable[Field]:
        """Iterate over all fields."""
        yield ExcelGeometryField(self)
        yield ExcelDataField(self)
